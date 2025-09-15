import json
import os
import re
from datetime import datetime
from helpers.common_utils import get_daily_count
from config import URLS
from pathlib import Path
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from pathlib import Path
from playwright.sync_api import Page, expect
import openpyxl

PRODUCT_FILE_PATH = Path("product_name.json")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 제품명 생성 함수
def generate_product_names():
    now = datetime.now()
    cnt = get_daily_count()
    date = now.strftime("%m%d")
    count = f"{cnt:02d}"
    prdname_kor = f"등록테스트_{date}_{count}"
    prdname_eng = f"TestProduct_{date}_{count}"
    return prdname_kor, prdname_eng
def generate_product_name(count:int):
    prdname_kor = f"배치 확인 제품_{count}"
    prdname_eng = f"AutoProduct_{count}"
    return prdname_kor, prdname_eng

# 제품 등록 후 json 파일에 제품 정보 업로드
def append_product_name(
    prdname_kor: str,
    prdname_eng: str,
    type_name: str,
    group: str,
    maker: str,
    order_rule: str,
    supplier: str,
    approve_rule: str,
    register: str,
    safety: int = 0,
    auto_order : int =0,
    order_flag : int = 0,
    stock_qty : int =0,   
    delivery_status : int=0, # 1: 발주 요청, 2: 발주 진행, 3: 배송 진행, 4: 수령 완료(운송장O), 5: 발주 취소, 6: 발주 실패, 7: 수령 완료(운송장X)
    
):
    
    try:
        with open(PRODUCT_FILE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            if not isinstance(data, list):
                data = []
    except (FileNotFoundError, json.JSONDecodeError):
        data = []

    data.append({
        "kor": prdname_kor,
        "eng": prdname_eng,
        "supplier": supplier,
        "order_rule": order_rule,
        "type": type_name,
        "group": group,
        "maker": maker,
        "safety" : safety,
        "auto_order": auto_order,
        "order_flag" : order_flag,
        "stock_qty" : stock_qty,
        "delivery_status" : delivery_status,
        "approve_rule" : approve_rule,
        "register": register
    })

    with open(PRODUCT_FILE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return prdname_kor, prdname_eng

# json에 저장된 제품정보 모두 불러오기
def get_all_product_names():
    try:
        with open(PRODUCT_FILE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

#json 파일에 최근 등록된 제품 정보 불러오기
def get_latest_product_name():
    all_names = get_all_product_names()
    if not all_names:
        raise ValueError("❌ 저장된 제품명이 없습니다.")
    return all_names[-1]

#json 파일 내 제품 삭제
def remove_products_from_json(deleted_names: list):
    """
    주어진 제품명 목록을 기반으로 JSON 파일에서 해당 제품들을 제거합니다.

    Args:
        deleted_names (list): 삭제할 제품명의 리스트.
    """
    try:
        with open(PRODUCT_FILE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        # 삭제할 제품명을 가진 항목 필터링
        updated_data = [item for item in data if item.get("kor") not in deleted_names]

        with open(PRODUCT_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(updated_data, f, ensure_ascii=False, indent=2)
        

    except Exception as e:
        print(f"[ERROR] JSON 파일 업데이트 중 오류 발생: {e}")

# 제품 등록 이후 해당 제품명 리스트 찾기
def verify_products_in_list(page, product_names: list[str], url: str,  table_column_index: int):
    page.goto(url)
    page.wait_for_selector(f"data-testid=input_search", timeout=10000)
    
    for name in product_names:
        page.fill("data-testid=input_search", name)
        page.wait_for_timeout(1000)
        page.click("data-testid=btn_search")
        page.wait_for_timeout(3000)

        rows = page.locator("table tbody tr")
        found = False
        for i in range(rows.count()):
            row = rows.nth(i)
            cell_text = row.locator(f"td:nth-child({table_column_index})").inner_text().strip()
            if name in cell_text:
                print(f"[PASS] {name} → '{url}'에서 확인됨")
                page.wait_for_timeout(1000)
                page.locator("data-testid=btn_reset").click()
                page.wait_for_timeout(1000)
                found = True
                break

        if not found:
            raise AssertionError(f"[FAIL] {name} → '{url}'에서 확인되지 않음")

def is_product_exist(page, product_names) -> bool:
    if isinstance(product_names, str):
        product_names = [product_names]

    all_exist = True
    for name in product_names:
        try:
            page.fill("data-testid=input_search", name)
            page.wait_for_timeout(1000)
            page.locator("data-testid=btn_search").click()
            page.wait_for_timeout(3000)

            # 검색 결과 로딩 대기
            page.locator("table tbody tr").first.wait_for(timeout=5000)

            rows = page.locator("table tbody tr")
            found = False
            for i in range(rows.count()):
                row_name = rows.nth(i).locator("td:nth-child(4) div.truncate").first.inner_text().strip()
                if name in row_name:
                    print(f"[PASS] '{name}' found in 제품 리스트")
                    found = True
                    page.locator("data-testid=btn_reset").click()
                    page.wait_for_timeout(1000)
                    break

            if not found:
                print(f"[FAIL] '{name}' not found in 제품 리스트")
                all_exist = False

        except PlaywrightTimeoutError:
            print(f"[FAIL] Timeout while searching for '{name}'")
            all_exist = False

    return all_exist

# 실제 등록된 리스트와 json 파일 비교 하여 업데이트
def sync_product_names_with_server(page):
    product_list = get_all_product_names()
    valid_list = []
    
    page.goto(URLS["bay_prdList"])

    for item in product_list:
        if is_product_exist(page, item["kor"]):
            valid_list.append(item)
        else:
            print(f"[삭제됨] 서버에 없는 제품명 제거: {item['kor']}")
            remove_products_from_json(item["kor"])  # JSON에서 제거

    return valid_list

#제품명 변경 후 json 파일 업데이트
def update_product_name(old_kor: str, new_kor: str):
    path = "product_name.json"
    if not os.path.exists(path):
        return

    with open(path, "r", encoding="utf-8") as f:
        products = json.load(f)

    for product in products:
        if product.get("kor") == old_kor:
            product["kor"] = new_kor
            break

    with open(path, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    print(f"[INFO] 제품명 업데이트 완료: {old_kor} → {new_kor}")

# 제품 수정 후 json 파일 업데이트
def update_product_flag(name_kor: str, **flags):
    """제품의 재고 및 플래그 값 업데이트"""
    path = "product_name.json"
    if not os.path.exists(path):
        return

    with open(path, "r", encoding="utf-8") as f:
        products = json.load(f)

    for product in products:
        json_name = product.get("kor", "").replace("\n", "").strip()
        target_name = name_kor.replace("\n", "").strip()

        if json_name == target_name:
            for key, value in flags.items():
                if value is not None:
                    product[key] = value
            updated = True
            break

    with open(path, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

# 저장된 제품명 목록 불러오기
def load_saved_product_names():
    path = "product_name.json"
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)
# 수정 후 수정 값 확인 (검색 포함)
def verify_product_update(page, product_name):
    name = product_name
    page.goto(URLS["bay_prdList"])
    page.wait_for_selector("data-testid=input_search", timeout=5000)
    page.fill("data-testid=input_search", name)
    page.wait_for_timeout(500)
    page.click("data-testid=btn_search")
    page.wait_for_timeout(3000)

    # 4열에 해당 제품명이 있는지 확인
    rows = page.locator("table tbody tr")
    found = False

    # Locator를 사용하여 각 행에 접근
    row_count = rows.count()
    for i in range(row_count):
        row = rows.nth(i)  # 각 행을 하나씩 가져오기
        product_name = row.locator("td:nth-child(4)").inner_text().strip().split("\n")[0]
        
        # 말줄임 처리된 텍스트를 처리하기 위해 title 속성 값 사용
        if not product_name:  # 만약 텍스트가 없다면
            product_name = row.locator("td:nth-child(4)").get_attribute("title").strip()

        # 공백을 제거하고 비교
        product_name = product_name.replace(" ", "").strip()
        name = name.replace(" ", "").strip()

        print(f"UI에서 노출되는 제품명: '{product_name}'")
        print(f"비교하는 제품명: '{name}'")
        
        if product_name == name:
            found = True
            print(f"수정한 제품명: {name}")
            break

    # 제품명이 없으면 실패 처리
    if not found:
        print(f"❌ 제품 관리 페이지에서 수정 확인 실패: {name}")
        return False  # 수정된 제품명이 UI에 반영되지 않으면 False 반환
    return True  # 모든 제품명이 일치하면 True 반환

# 특정 제품의 현 재고량 찾기
def get_product_stock(page, product_name):
    from config import URLS
    page.goto(URLS["bay_stock"])
    page.wait_for_selector("data-testid=input_search", timeout=5000)
    page.fill("input[placeholder='제품명 검색']", product_name)
    page.wait_for_timeout(500)
    page.click("data-testid=btn_search")
    page.wait_for_timeout(1000)

    rows = page.locator("table tbody tr")
    for i in range(rows.count()):
        name_cell = rows.nth(i).locator("td:nth-child(5)").inner_text().strip()
        if product_name in name_cell:
            stock_text = rows.nth(i).locator("td:nth-child(3)").inner_text().strip()
            return int(stock_text) if stock_text.isdigit() else 0

    raise Exception(f"❌ 재고관리에서 제품 '{product_name}'을 찾을 수 없음")

# 업체 중복값 확인을 위한 정보 불러오기
def is_duplicate_supplier_from_product_file(manager: str, contact: str) -> bool:
    try:
        with open(PRODUCT_FILE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return False

    for item in data:
        if (
            item.get("supplier") == "자동화 업체명" and
            item.get("manager") == manager and
            item.get("contact") == contact
        ):
            return True
    return False

# 등록한 업체 정보 값 찾기(페이지네이션 포함)
def find_supplier_in_paginated_list(page, supplier: str, manager: str, contact: str, memo : str) -> bool:
    # 검색
    page.locator("data-testid=input_search").fill(supplier)
    page.wait_for_timeout(1000)
    page.click("data-testid=btn_search")
    page.wait_for_timeout(1000)
    rows = page.locator("table tbody tr")
    for i in range(rows.count()):
        row = rows.nth(i)
        row_text = row.inner_text()
        if supplier in row_text and manager in row_text and contact and memo in row_text:
            return True
    
# 출고 테스트를 위한 제품 0 아닌 제품 찾기
def get_outflow_target_products():
    with open("product_name.json", "r", encoding="utf-8") as f:
        products = json.load(f)

    # stock 값이 0이 아닌 제품만 필터링
    eligible = [p for p in products if p.get("stock", 0) != 0]
    return eligible

# 제품 등록 시 드롭다운 선택하는 공통 함수
def select_from_dropdown(page, trigger_id: str, search_id: str, item_id: str, keyword: str) -> str:
    page.locator(f"[data-testid='{trigger_id}']").last.click()
    page.wait_for_timeout(1000)
    page.fill(f"[data-testid='{search_id}']", keyword)
    page.wait_for_timeout(1000)  
    page.locator(f"[data-testid='{item_id}']", has_text=keyword).click()
    page.wait_for_timeout(1000)
    return keyword

def check_rule_for_products(page, products, col_index: int, expected_key: str, label: str):
    
    for product in products:
        page.locator("data-testid=input_search").fill(product["kor"])  # 제품명 검색
        page.wait_for_timeout(500)
        page.locator("data-testid=btn_search").click()
        page.wait_for_timeout(1000)

        rows = page.locator("table tbody tr")
        first_row = rows.nth(0)
        rule_cell = first_row.locator(f"td:nth-child({col_index})")  # 지정한 열 가져오기
        rule_text = rule_cell.inner_text().strip()

        # 기대값: json key 기반 vs 고정 값
        expected_value = product[expected_key] if expected_key else "자동 승인"
        
        assert rule_text == expected_value, \
            f"{label} 불일치 → 기대: {expected_value}, 실제: {rule_text}"

        print(f"✅ {label} 확인 완료: {product['kor']} ({rule_text})")
        page.wait_for_timeout(1000)

def edit_approval_rules_and_check(page, products):
    for product in products:
        product_name = product["kor"]
        register_type = product.get("register")

        # register 값에 따라 승인 규칙 결정
        if register_type == "excel":
            approval_rule = "승인규칙_1명"
        elif register_type == "manual":
            approval_rule = "승인규칙_n명"
        else:
            raise ValueError(f"❌ 알 수 없는 register 값: {register_type}")

        # 1. 제품 검색
        page.locator("data-testid=input_search").fill(product_name)
        page.wait_for_timeout(500)
        page.locator("data-testid=btn_search").click()
        page.wait_for_timeout(2000)

        rows = page.locator("table tbody tr")
        first_row = rows.nth(0)
        edit_button = first_row.locator("td:last-child >> text=수정")
        edit_button.click()
        page.wait_for_timeout(2000)

        # 2. 승인 규칙 드롭다운 선택
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(500)
        page.locator("data-testid=drop_approval_trigger").click()
        page.wait_for_timeout(500)
        page.locator("data-testid=drop_approval_search").fill(approval_rule)
        page.wait_for_timeout(500)
        page.locator("data-testid=drop_approval_item", has_text=approval_rule).click()
        page.wait_for_timeout(500)

        # 3. 저장 → 토스트 확인
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(500)
        page.locator("data-testid=btn_save").click()
        expect(page.locator("data-testid=txt_edit")).to_have_text("제품을 수정하시겠습니까?", timeout=3000)
        page.locator("data-testid=btn_confirm").click()
        expect(page.locator("data-testid=toast_edit")).to_be_visible(timeout=3000)
        page.wait_for_timeout(1000)

        # 4. 다시 제품 리스트에서 검색하여 승인 규칙 확인
        page.locator("data-testid=input_search").fill(product_name)
        page.wait_for_timeout(500)
        page.locator("data-testid=btn_search").click()
        page.wait_for_timeout(2000)

        rows = page.locator("table tbody tr")
        first_row = rows.nth(0)
        approval_cell = first_row.locator("td:nth-child(12)")  # 승인 규칙명 12열
        approval_text = approval_cell.inner_text().strip()

        assert approval_text == approval_rule, \
            f"승인 규칙 적용 실패 → 기대: {approval_rule}, 실제: {approval_text}"

        print(f"✅ {product_name} ({register_type}) → {approval_rule} 적용 확인 완료")

# 엑셀 파일 업로드 + 업로드 성공 확인 + 엑셀 행 수 vs UI 행 수 비교까지 한 번에 처리
def upload_and_verify_excel(page: Page, file_path: str, table_selector: str = "table tbody tr"):
    # 업로드
    page.wait_for_selector("data-testid=btn_excel", timeout=5000)
    page.locator("data-testid=btn_excel").hover()
    page.wait_for_selector("data-testid=btn_upload", timeout=5000)

    page.set_input_files("input[type='file']", file_path)
    print(f"📂 업로드 요청: {file_path}")

    page.wait_for_selector("data-testid=col_type", timeout=10000)
    expect(page.locator("data-testid=btn_save")).to_be_disabled(timeout=3000)
    print(f"⬆️ 업로드 완료: {file_path}")

    # 엑셀 파일 로드해서 행 수 확인
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    excel_rows = sum(1 for row in sheet.iter_rows(min_row=2, values_only=True) if any(row))

    # UI 테이블 행 수 확인
    ui_rows = page.locator(table_selector).count()
    assert excel_rows == ui_rows, f"엑셀 {excel_rows}행 vs UI {ui_rows}행 불일치"
    print(f"✅ {file_path}: 엑셀 {excel_rows}행 = UI {ui_rows}행 일치")

    return headers, excel_rows

# 엑셀 업로드를 위한 엑셀 데이터 업데이트
def update_product_names(file_path="data/success.xlsx"):
    now = datetime.now()
    date = now.strftime("%m%d")

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    col_kor, col_eng = 5, 6  # E열=제품명, F열=제품명(영문)

    for i, row in enumerate(sheet.iter_rows(min_row=2, max_col=col_eng, values_only=False)):
        # ✅ 1열(A열)에 값이 있을 때만 처리
        if row[0].value:
            cnt = get_daily_count()  # 날짜별 카운터 가져오기
            count = f"{cnt:02d}"

            row[col_kor-1].value = f"엑셀업로드_{date}_{count}"
            row[col_eng-1].value = f"upload_product_{date}_{count}"

    workbook.save(file_path)
    print(f"✅ 제품명 업데이트 완료 ({date}, 마지막 번호 {count})")